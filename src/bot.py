import discord
from discord.ext import commands
import random
import openpyxl
import asyncio
import csv
import os

# === Load Excel Sheet and Groups ===
wb = openpyxl.load_workbook("../words/Word-Groups-with-meanings.xlsx", data_only=True)
group_names = wb.sheetnames  # E.g., ["Group1", "Group2", ...]

groups = {}
for sheet_name in group_names:
    sheet = wb[sheet_name]
    items = [
        (row[0], row[1])
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[0] and row[1]
    ]
    groups[sheet_name.lower()] = items

all_word_meanings = [
    (word, meaning)
    for group in groups.values()
    for word, meaning in group
]

# === Load fallback sheet "SheetB" ===
# Replace 'SheetB' with the actual fallback sheet name if different
fallback_sheet_name = "../words/words.xlsx"
try:
    sheet_B = wb[fallback_sheet_name]
    sheet_B_word_meanings = [
        (row[0], row[1]) for row in sheet_B.iter_rows(min_row=2, values_only=True) if row[0] and row[1]
    ]
except KeyError:
    print(f"Warning: Fallback sheet '{fallback_sheet_name}' not found. Using all groups for fallback.")
    sheet_B_word_meanings = all_word_meanings

# === Bot Setup ===
intents = discord.Intents.default()
intents.message_content = True
intents.members = True
bot = commands.Bot(command_prefix="!", intents=intents)
bot.remove_command("help")

# Game state
ratings_file = "ratings.csv"
ratings = {}  # user_id -> rating
BUTTON_OPTIONS = ["🇦", "🇧", "🇨", "🇩"]

# === Load ratings from CSV ===
if os.path.exists(ratings_file):
    with open(ratings_file, newline='') as f:
        reader = csv.reader(f)
        for row in reader:
            try:
                ratings[int(row[0])] = float(row[1])
            except:
                continue
else:
    open(ratings_file, "w").close()

# === Save ratings to CSV ===
def save_ratings():
    with open(ratings_file, "w", newline="") as f:
        writer = csv.writer(f)
        for user_id, rating in ratings.items():
            writer.writerow([user_id, round(rating, 2)])

# === Elo Rating ===
def update_ratings(players_scores):
    K = 32
    sorted_players = sorted(players_scores.items(), key=lambda x: x[1], reverse=True)
    winners = [u for u, s in sorted_players if s == sorted_players[0][1]]
    losers = [u for u, s in sorted_players if s < sorted_players[0][1]]

    for winner in winners:
        ratings[winner.id] = ratings.get(winner.id, 1000)
    for loser in losers:
        ratings[loser.id] = ratings.get(loser.id, 1000)

    for winner in winners:
        for loser in losers:
            Ra = ratings[winner.id]
            Rb = ratings[loser.id]
            Ea = 1 / (1 + 10 ** ((Rb - Ra) / 400))
            Eb = 1 / (1 + 10 ** ((Ra - Rb) / 400))
            ratings[winner.id] = Ra + K * (1 - Ea)
            ratings[loser.id] = Rb + K * (0 - Eb)
    save_ratings()

# === Quiz Button View ===
class QuizButton(discord.ui.Button):
    def __init__(self, label, choice, view):
        super().__init__(label=label, style=discord.ButtonStyle.primary)
        self.choice = choice
        self.view_ref = view

    async def callback(self, interaction: discord.Interaction):
        if interaction.user in self.view_ref.answered:
            await interaction.response.send_message("❌ You already answered!", ephemeral=True)
            return
        self.view_ref.answered[interaction.user] = self.choice
        await interaction.response.send_message(f"✅ You chose: **{self.choice}**", ephemeral=True)

class QuizView(discord.ui.View):
    def __init__(self, options):
        super().__init__(timeout=None)
        self.answered = {}  # user -> choice
        for i, option in enumerate(options):
            self.add_item(QuizButton(BUTTON_OPTIONS[i], option, self))

# === Start Quiz Command ===
@bot.command()
async def startquiz(ctx, rounds: int = 5, round_length: int = 10, group: str = None):
    """
    Starts a button-based word quiz (optionally for a specific group).
    Usage: !startquiz [rounds] [round_length] [groupname]
    """
    players_scores = {}
    used_words = set()  # Track words already asked in this quiz

    chosen_group = None
    if group:
        group_key = group.lower()
        chosen_group = groups.get(group_key)
        if not chosen_group:
            await ctx.send(f"❌ Group `{group}` not found! Valid groups: " + ", ".join(groups.keys()))
            return

    await ctx.send(
        f"🎮 Starting a quiz with {rounds} rounds, {round_length}s per round!"
        + (f" Using group '{group.title()}' only for question words." if chosen_group else "")
        + (f" Using fallback sheet '{fallback_sheet_name}'." if not chosen_group else "")
    )

    for round_num in range(1, rounds + 1):
        # Determine source for questions:
        if chosen_group:
            word_list_source = chosen_group
        else:
            word_list_source = sheet_B_word_meanings

        # Select a word not already used (if possible)
        available_words = [wm for wm in word_list_source if wm[0] not in used_words]
        if not available_words:
            # All words used, fallback to full list (allow repeats)
            available_words = word_list_source

        word, correct_meaning = random.choice(available_words)
        used_words.add(word)

        # Distractor options always from all groups:
        wrong_options_pool = [m for _, m in all_word_meanings if m != correct_meaning]
        wrong_options = random.sample(wrong_options_pool, 3) if len(wrong_options_pool) >= 3 else wrong_options_pool
        options = wrong_options + [correct_meaning]
        random.shuffle(options)

        description = "\n".join([f"{BUTTON_OPTIONS[i]} - {options[i]}" for i in range(4)])
        embed = discord.Embed(
            title=f"Round {round_num}/{rounds}",
            description=f"📖 **{word.upper()}**\n\n{description}\n\n⏳ Time left: {round_length}s",
            color=discord.Colour.blue()
        )

        view = QuizView(options)
        msg = await ctx.send(embed=embed, view=view)

        # Countdown timer
        for t in range(round_length, 0, -1):
            try:
                timer_embed = discord.Embed(
                    title=f"Round {round_num}/{rounds}",
                    description=f"📖 **{word.upper()}**\n\n{description}\n\n⏳ **Time left: {t}s** ⏳",
                    color=discord.Colour.blue()
                )
                await msg.edit(embed=timer_embed)
                await asyncio.sleep(1)
            except discord.NotFound:
                break

        reacted_users = view.answered

        embed = discord.Embed(title="Round Results", color=discord.Colour.green())
        embed.add_field(name="Correct Answer", value=f"✅ {correct_meaning}", inline=False)
        if reacted_users:
            for user, choice in reacted_users.items():
                score = 1 if choice == correct_meaning else 0
                players_scores[user] = players_scores.get(user, 0) + score
                embed.add_field(name=user.name, value=f"Chose: {choice} | +{score} points", inline=False)
        else:
            embed.add_field(name="No answers", value="😢 No one answered!", inline=False)
        await ctx.send(embed=embed)
        await asyncio.sleep(2)

    embed = discord.Embed(title="🏆 Final Scores 🏆", color=discord.Colour.gold())
    if players_scores:
        sorted_scores = sorted(players_scores.items(), key=lambda x: x[1], reverse=True)
        for user, score in sorted_scores:
            embed.add_field(name=user.name, value=f"{score} points", inline=False)
        await ctx.send(embed=embed)
        update_ratings(players_scores)
        rating_embed = discord.Embed(title="📊 Updated Ratings 📊", color=discord.Colour.purple())
        for user in players_scores:
            rating_embed.add_field(name=user.name, value=f"{round(ratings[user.id])}", inline=False)
        await ctx.send(embed=rating_embed)
    else:
        await ctx.send("No one scored any points! 😅")

# === List Groups Command ===
@bot.command()
async def listgroups(ctx):
    embed = discord.Embed(title="Available Word Groups", color=discord.Colour.blue())
    embed.description = "\n".join([f"• {name.title()}" for name in groups.keys()])
    await ctx.send(embed=embed)

# === Leaderboard Command ===
@bot.command()
async def leaderboard(ctx):
    if not ratings:
        await ctx.send("No ratings yet! 😅")
        return
    top = sorted(ratings.items(), key=lambda x: x[1], reverse=True)[:10]
    embed = discord.Embed(title="🏅 Leaderboard", color=discord.Colour.gold())
    for user_id, rating in top:
        try:
            user = await bot.fetch_user(user_id)
            embed.add_field(name=user.name, value=f"{round(rating)}", inline=False)
        except:
            continue
    await ctx.send(embed=embed)

# === Help Command ===
@bot.command()
async def quizhelp(ctx):
    embed = discord.Embed(title="📖 GRE Quiz Bot Help", color=discord.Colour.blue())
    embed.add_field(
        name="!startquiz [rounds] [round_length] [groupname]",
        value=(
            "Start a button-based word quiz.\n"
            "• rounds: number of questions (default 5)\n"
            "• round_length: seconds per round (default 10)\n"
            "• groupname: quiz only from this word group (see !listgroups)"
        ),
        inline=False
    )
    embed.add_field(
        name="!leaderboard",
        value="Show top 10 players and their Elo ratings.",
        inline=False
    )
    embed.add_field(
        name="!listgroups",
        value="Show all available word groups (sheet names in the Excel file).",
        inline=False
    )
    embed.add_field(
        name="!quizhelp",
        value="Show this help message.",
        inline=False
    )
    embed.set_footer(text="Have fun quizzing! 🎓")
    await ctx.send(embed=embed)

# === Run Bot ===
# bot.run("YOUR_BOT_TOKEN")

