import openpyxl
from nltk.corpus import wordnet as wn
import nltk

nltk.download('wordnet')

def get_first_wordnet_meaning(word):
    synsets = wn.synsets(str(word))
    if not synsets:
        return 'Meaning not found'
    first_def = synsets[0].definition()
    return first_def.strip()

input_filename = "Word-Groups.xlsx"
output_filename = "Word-Groups-with-meanings.xlsx"

wb = openpyxl.load_workbook(input_filename)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"Processing sheet: {sheet_name} ({sheet.max_row - 1} words)")

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1)
        word = cell.value
        if word is None or str(word).strip() == "":
            continue
        meaning = get_first_wordnet_meaning(word)
        # Write meaning in col 2, creating it if not present
        sheet.cell(row=row, column=2, value=meaning)

wb.save(output_filename)
print(f"All group meanings saved to {output_filename}!")
